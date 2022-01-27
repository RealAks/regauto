# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
def getColumnHeaders(row_detail, prev_Col_Header, colStart):
    #print(len(row_detail)," ", colStart)
    for i in range(len(row_detail)):
        if i >= colStart:
            #print(row_detail[i])
            if row_detail[i] is not None:
                #print("this row is a column bcos of ", row_detail[i].split())
                return "Y"
    return "N"

#print("saving done")


# %%
# Initial Configuration
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np 
import win32com.client
from win32com.client import Dispatch


#filename = "Book1.xlsx"
#path     = "D:/Akhil/Work/DOTS/RegAutomation/Akhil/input/20210621"
mode     = "N"  # N(new), M (modified)

#print("After specifying File details")


# %%
# Initialize Variables
rowidx = 0
colidx = 0
colStartidx = 2
columnHeaders = []
columnHeadersList = []
isColHeaderPresent = 'N'
reportBlockList = []
lastColRow = 0
nextRow = 0
reportListBlock = []
offset = 0
eachRow = []
isResetColStrt = "N"
isColContinous = "N"
referenceSheetNames = []
sheetName = ""
#print("after initializing variables")


# %%
#Generate Preview File --> With LineItems and Derived Line Items
#print(path+"/Inp-"+filename)
if mode == "N":
#    wb = openpyxl.load_workbook(path+"/Inp-"+filename)
#    ws = wb.worksheets[0]
    xl = win32com.client.Dispatch('Excel.Application')
    wb = xl.Workbooks.Open(r'C:\Users\i31927\Desktop\MLProject\ML-ADDIN\test.xlsm')
    ws = wb.Worksheets('INPUT_DRAFT')
    #print("after opening files")
    #Delete any existing Section / Blocks
    rowidx = 0
    reportBlockList = []
    myrow = range(1,1000)
    for row in myrow:
        rowidx+=1
        eachRow = []
        eachRowColour = []
        firstWordInRow = ""
        
        for cell in myrow:
            eachRow.append(cell.value)
            eachRowColour.append(cell.fill.start_color.index)
            #print("firstWordInRow is ", firstWordInRow, '+', cell.value)
            if len(firstWordInRow) < 1 and cell.value is not None:
                firstWordInRow = cell.value
                #print("firstWordInRow is ", firstWordInRow)
            #print(cell.value)
            
        #print("Gonna Check ",firstWordInRow)
        if len(firstWordInRow) > 0:
            #print("So it is ",firstWordInRow[0])
            if firstWordInRow[0] == '{':
                #print("Adding row ", rowidx)
                reportBlockList.append(rowidx)

    #print("out of the crazy loop")
    for blk in range(len(reportBlockList)):
        #print("deleting row no ", reportBlockList[blk]-blk)
        ws.delete_rows(reportBlockList[blk]-blk)
        
    #Iterate Rows and identify Column Headers
    rowidx = 0
    reportBlockList = []
    for row in ws.iter_rows():
        rowidx+=1
        eachRow = []
        firstVal = ""
        for cell in row:
            eachRow.append(cell.value)
            if len(firstVal) < 1 and cell.value is not None:
                firstVal = cell.value 
             

        ## Need to fix the logic for identifying the Column Start Index 
        #Check if Row is a Column Header
        columnHeaders = []
        isColHeaderPresent = getColumnHeaders(eachRow, columnHeadersList, colStartidx)
        if isColHeaderPresent == 'Y':
            #print("came here")
            if isColContinous == 'N':    ## reset ColumnHeaders if new list of Column Headers
                #print("came here 2", rowidx)
                columnHeadersList = [] 
                reportBlockList.append(rowidx)
                lastColRow = rowidx
                nextRow = 0

            columnHeadersList.append(eachRow)
            isColContinous = 'Y'
            nextRow+=1
        else:
            isColContinous = 'N'
            #print("How is ", rowidx, " = ", lastColRow, " + ", nextRow," and firstVal be ", firstVal)
            if rowidx == lastColRow+nextRow and len(firstVal) == 0:
                #print("Column Row followed by Empty Row: ", lastColRow)
                reportListBlock.append(lastColRow)
    
    #Insert Report Sections and Blocks
    #print("here after identifying report blocks/sections")
    ws.insert_rows(0, 9)
    ws.cell(row = 2, column = 1).value = "{ReportSection:Start}"
    ws.cell(row = 3, column = 1).value = "{SectionType:NonGroup}"
    ws.cell(row = 4, column = 2).value = "{Image:banklogo.jpg}"
    ws.cell(row = 5, column = 1).value = "{ReportSection:End}"
    ws.cell(row = 7, column = 2).value = "{ReportSection:Start}"
    ws.cell(row = 8, column = 2).value = "{SectionType:NonGroup}"
    ws.cell(row = 9, column = 2).value = "{ReportBlock:Start}"
    ws.cell(row = 10, column = 2).value = "{BlockType:Grid}"
    offset = 9
    for blk in range(len(reportBlockList)):
        if blk > 0:
            rowidx = reportBlockList[blk] + offset
            ws.insert_rows(rowidx,3)
            ws.cell(row = rowidx, column = 1).value = "{ReportBlock:End}"
            rowidx+=1
            ws.cell(row = rowidx, column = 1).value = "{ReportBlock:Start}"
            rowidx+=1
            if reportBlockList[blk] in reportListBlock:
                ws.cell(row = rowidx, column = 1).value = "{BlockType:List}"
            else:
                ws.cell(row = rowidx, column = 1).value = "{BlockType:Grid}"

            offset +=3
        #print(reportBlockList[blk])

    rowidx = ws.max_row+1
    ws.cell(row = rowidx, column = 1).value = "{ReportBlock:End}"
    rowidx +=1
    ws.cell(row = rowidx, column = 1).value = "{ReportSection:End}"

#print("processing done") 

wb.save(r'C:\Users\i31927\Desktop\MLProject\ML-ADDIN\test.xlsm')


# %%



