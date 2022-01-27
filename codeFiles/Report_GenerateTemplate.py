# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
def getSectionHeader(row, isColHead, colStart):
    splitString = []
    sectionH = ""
    #print("Inside function getSection", row[0], ' $&$ ', row[1])
    #for i in range(len(row)):
    i = 0
    isCapsChecked = "N"
    while i < colStart:
        #print("while we check ", row[i])
        if row[i] is not None:
            splitString.append(row[i])
            sectionH = sectionH + row[i]
        i+=1

    if isColHead == 'Y':
        #print("section bcos of col header ", sectionH)
        return sectionH
    
    #print("Col starts from ",colStart)
    #print("length is ", len(splitString))
    for j in range(len(splitString)):
        if splitString[j].upper().split()=="PART" or splitString[j].upper().split()=="SECTION":
            #print("Section bcos of Part/Section", sectionH)
            return sectionH
        #print("String being checked", splitString[j])
        if isCapsChecked == "N" and splitString[j].split() != "":
            isCapsChecked = "Y"
            if splitString[j].find(".") != -1:
                #print("After checking", splitString[j].split(".")[0])
                res = splitString[j].split(".")[0]
                res.replace(".", "")
                #print("section being checked: ", res)
                #if res.isupper() == True or res.isdigit() == True:
                if res.isupper() ==True and res.isalpha() == True:
                    #print("Section bcos of Capital letter")
                    return sectionH
    return ""      

def getColumnHeaders(row_detail, prev_Col_Header, colStart):
    #print(len(row_detail)," ", colStart)
    for i in range(len(row_detail)):
        if i >= colStart:
            #print(row_detail[i])
            if row_detail[i] is not None and row_detail[i] not in ("&&LI&&", "&&DLI&&"):
                if len(row_detail[i])>0:
                    if row_detail[i][0] == "=":
                        #print("ignore.. is a formulae")
                        continue;                
                #print("this row is a column bcos of ", row_detail[i].split())
                return "Y"
    return "N"

def formatStringForLIDesc(string):
    start = 0
    if string[0].isdigit() == True:
        start = string.find(" ") + 1
    res = string[start:]
    res = res.replace(".", "")
    res = res.replace("(", "")
    res = res.replace("0", "")
    #print("formated result: ", res)
    return res            


#print("saving done")


# %%
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import pandas as pd
import win32com.client
from win32com.client import Dispatch

#filename = "Book1.xlsx"
mdfilename ="metadata.csv"
#path     = "D:/Akhil/Work/DOTS/RegAutomation/Akhil/input/20210621"
mode     = "N"  # N(new), M (modified)

#print("After specifying File details")


# %%
#Initiate Variables
rowidx = 0
colidx = 0
colStartidx = 2
columnHeaders = []
columnHeadersList = []
isColHeaderPresent = 'N'
isSectionHeader = 'N'
sectionHeader = ""
prevSectionHeader = ""
firstWordInRow = ""
eachRow = []
eachRowColour = []
allRows = []
liMetadata = []
liMetadataList = []
lineItemId = 0
sectionId = 0
lineItemDesc = ""
rowDesc = ""
isResetColStrt = "N"
isColContinous = "N"
isRowReportConfig = "N"
referenceSheetNames = []
sheetName = ""
#print("After Initiating Variables")


# %%
#Second Part of the Proc - Generate Final File
if mode == "N":
  #  wb = openpyxl.load_workbook(path+"/Inp3-"+filename)
  #  ws = wb.worksheets[0]
  xl = win32com.client.Dispatch('Excel.Application')
  wb = xl.Workbooks.Open(r'C:\Users\i31927\Desktop\MLProject\ML-ADDIN\test.xlsm')
  ws = wb.Worksheets('INPUT_DRAFT')

   ## Add Logic to Traverse through multiple Sheets in an Excel
    
    nvl = lambda x, y: x if x is not None else y
    for row in ws.iter_rows():
         
        rowidx+=1
        eachRow = []
        eachRowColour = []
        firstWordInRow = ""
        for cell in row:
            eachRow.append(cell.value)
            eachRowColour.append(cell.fill.start_color.index)
            #print(cell.value)
            if len(firstWordInRow) < 1 and cell.value is not None:
                firstWordInRow = cell.value 
                #print("firstWordInRow is ", firstWordInRow)
            #print(cell.value)
            ## Change the below Logic to Traverse through multiple Report Sections & Report Blocks
        if len(firstWordInRow) > 0:
            if firstWordInRow[0] == '{':
                #print("In Here")
                columnHeadersList = []
                continue;
        ## Need to fix the logic for identifying the Column Start Index 
        #Check if Row is a Column Header
        columnHeaders = []
        isColHeaderPresent = getColumnHeaders(eachRow, columnHeadersList, colStartidx)
        if isColHeaderPresent == 'Y':
            #print("came here")
            if isColContinous == 'N':    ## reset ColumnHeaders if new list of Column Headers
                #print("came here 2")
                columnHeadersList = [] 
                isResetColStrt = "Y"
            columnHeadersList.append(eachRow)
            isColContinous = 'Y'
        else:
            isColContinous = 'N'
        #print("Is Column? ", isColHeaderPresent)
        #print("Col Headers are ", len(columnHeadersList))
        #Check if Row is a separate Section
        sectionHeader = getSectionHeader(eachRow, isColHeaderPresent, colStartidx)
        #print("Section Header returned is ", sectionHeader, len(sectionHeader.split()))
        #print("Also ", prevSectionHeader)
        if len(sectionHeader.split()) > 0 and sectionHeader != prevSectionHeader:
            #print("inside if")
            prevSectionHeader = sectionHeader
            isSectionHeader = 'Y'
            sectionId+=1
        else:
            sectionHeader = prevSectionHeader
            isSectionHeader = 'N'
        #print("Is Section?", isSectionHeader)
        
        rowDesc = ""
        for eachCell in range(len(eachRow)):
            
            #print("Col ", eachCell, " is ", eachRow[eachCell], " with color ", eachRowColour[eachCell])
            if eachRow[eachCell] == "&&LI&&":
                lineItemId+=1
                columnHeaders = ""
                for eachCol in range(len(columnHeadersList)):
                    columnHeaders = columnHeaders + " " + nvl(columnHeadersList[eachCol][eachCell], "")
                lineItemDesc = formatStringForLIDesc(rowDesc) + formatStringForLIDesc(columnHeaders)
                ws.cell(row = rowidx, column = eachCell+1).value = "{Query1:CLIV:LI="+str(lineItemId)+",MD(RLA=Y:DD=Y:LIA=Y)}" 
                ws.cell(row = rowidx, column= eachCell+1).fill = PatternFill(start_color= '00000000', end_color='00000000', fill_type= None) 
                liMetadata.append(sectionId)
                liMetadata.append(sectionHeader)
                liMetadata.append(lineItemId)
                liMetadata.append(lineItemDesc)
                liMetadata.append("N")
                liMetadataList.append(liMetadata)
                liMetadata = []
                
            elif eachRow[eachCell] == "&&DLI&&":
                
                lineItemId+=1
                columnHeaders = ""
                for eachCol in range(len(columnHeadersList)):
                    columnHeaders = columnHeaders + " " + nvl(columnHeadersList[eachCol][eachCell], "")
                lineItemDesc = formatStringForLIDesc(rowDesc) + formatStringForLIDesc(columnHeaders)
                ws.cell(row = rowidx, column = eachCell+1).value = "Enter Formulae for DLI" 
                ws.cell(row = rowidx, column= eachCell+1).fill = PatternFill(start_color= '00000000', end_color='00000000', fill_type= None) 
                liMetadata.append(sectionId)
                liMetadata.append(sectionHeader)
                liMetadata.append(lineItemId)
                liMetadata.append(lineItemDesc)
                liMetadata.append("Y")
                liMetadataList.append(liMetadata)
                liMetadata = []

            elif eachRow[eachCell] is None:
                pass   
            elif len(eachRow[eachCell]) > 0:
                 if eachRow[eachCell][0] == "=":
                     lineItemId+=1
                     columnHeaders = ""
                     for eachCol in range(len(columnHeadersList)):
                         columnHeaders = columnHeaders + " " + nvl(columnHeadersList[eachCol][eachCell], "")
                     lineItemDesc = formatStringForLIDesc(rowDesc) + formatStringForLIDesc(columnHeaders)
                    
                     ws.cell(row = rowidx, column= eachCell+1).fill = PatternFill(start_color= '00000000', end_color='00000000', fill_type= None) 
                     liMetadata.append(sectionId)
                     liMetadata.append(sectionHeader)
                     liMetadata.append(lineItemId)
                     liMetadata.append(lineItemDesc)
                     liMetadata.append("Y")
                     liMetadataList.append(liMetadata)
                     liMetadata = []
                 else:
                     rowDesc = rowDesc + eachRow[eachCell]
                    
        allRows.append(eachRow)
        #rowidx+=1
        isColHeaderPresent = "N"
        isSectionHeader = "N"
#print("processing done") 

## Add additional Sheets (move them as Procs/Functions)
wbref = openpyxl.load_workbook(path+"/ReferenceSheet.xlsx")
#print(wbref.sheetnames)
referenceSheetNames = wbref.sheetnames

for i in range(len(referenceSheetNames)):
    
    sheetName = referenceSheetNames[i]
    #print(sheetName)
    wsref = wbref[sheetName]
    #print("after opening SheetName")
    
    ws1 = wb.create_sheet()
    ws1.title = sheetName
    #ws1.title = 'Prompt'
    #wb.create_sheet['Query1']
    #wb.create_sheet['QualitativeComments']
    #wb.create_sheet['Query2']
    #wb.create_sheet['Query3']
    #wb.create_sheet['Query4']
    #print("new sheet created")
    #ws1 = wb["Prompt"]
    #ws1.cell(row = rowidx, column = eachCell+1).value = "Start Printing"

    rowidx = 0
    colidx=0
    #ws1.font = Font(bold = True)
    for row in wsref.iter_rows():
        rowidx+=1
        colidx = 0
        for cell in row:
            colidx+=1
            #print(cell.value)
            ws1.cell(row = rowidx, column = colidx).value = cell.value
    #ws1.font = Font(bold = True)
    #print("lets see")



wb.save(r'C:\Users\i31927\Desktop\MLProject\ML-ADDIN\test.xlsm')
df = pd.DataFrame(liMetadataList, columns = ['SectionID', 'SectionName', 'LineItemId', 'LineItemDesc', 'IsDerivedLineItem'] )
df.to_csv(path+'/LI_Metadata.csv', index = None, header = True)                    
#print("all done")   


# %%



