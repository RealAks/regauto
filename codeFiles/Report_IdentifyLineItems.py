# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
def getColumnHeaders(row_detail, prev_Col_Header, colStart):
    #print(len(row_detail)," ", colStart)
    for i in range(len(row_detail)):
        if i >= colStart:
            #print(row_detail[i])
            if row_detail[i] is not None:
                #print("this row is a column bcos of ", row_detail[i].split())
                if len(row_detail[i]) > 0:
                    return "Y"
    return "N"
def formatStringForLIDesc(string):
    start = 0
    if string[0].isdigit() == True:
        start = string.find(" ") + 1
    res = string[start:]
    res = res.replace(".", "")
    res = res.replace("(", "")
    res = res.replace("0", "")
    #print("formated result: ", res)
    return res    
#print("saving done")


# %%
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import pandas as pd
import win32com.client
from win32com.client import Dispatch

#filename = "Book1.xlsx"
#path     = "D:/Akhil/Work/DOTS/RegAutomation/Akhil/input/20210621"
mode     = "N"  # N(new), M (modified)

#print("After specifying File details")


# %%
# Initialize Variables
rowidx = 0
colidx = 0
colStartidx = 2
firstWordInRow = ""
columnHeaders = []
columnHeadersList = []
isColHeaderPresent = 'N'
eachRow = []
eachRowColour = []
rowDesc = ""
isResetColStrt = "N"
isColContinous = "N"
isRowReportConfig = "N"
isListSection = "N"
lineItemDesc = ""
referenceSheetNames = []
sheetName = ""
#print("after initializing variables")


# %%
#Generate Preview File --> With LineItems and Derived Line Items
if mode == "N":
    xl = win32com.client.Dispatch('Excel.Application')
    wb = xl.Workbooks.Open(r'C:\Users\i31927\Desktop\MLProject\ML-ADDIN\test.xlsm')
    ws = wb.Worksheets('INPUT_DRAFT')
    #print("after opening files")

    ## Add Logic to Traverse through multiple Sheets in an Excel
    
    nvl = lambda x, y: x if x is not None else y
    
    rowidx = 0
    for row in ws.iter_rows():
        rowidx+=1
        colidx = 0
        for cell in row:
            colidx+=1
            #print(cell.value)
            if cell.value in ("&&LI&&","&&DLI&&"):
                #print("updating Cell ", rowidx, "x", colidx)
                ws.cell(row = rowidx, column=colidx).value = None
                ws.cell(row = rowidx, column = colidx).fill = PatternFill(start_color= '00000000', end_color='00000000', fill_type= None) 

    
    rowidx = 0
    colidx = 0
    #print("After reset")
    for row in ws.iter_rows():
        #print("start")
  
        rowidx+=1
        eachRow = []
        eachRowColour = []
        firstWordInRow = ""
        for cell in row:
            eachRow.append(cell.value)
            eachRowColour.append(cell.fill.start_color.index)
            #print("firstWordInRow is ", firstWordInRow, '+', cell.value)
            if len(firstWordInRow) < 1 and cell.value is not None:
                firstWordInRow = cell.value 
                #print("firstWordInRow is ", firstWordInRow)
            #print(cell.value)
            ## Change the below Logic to Traverse through multiple Report Sections & Report Blocks
        if len(firstWordInRow) > 0:
            if firstWordInRow[0] == '{':
                #print("In Here")
                columnHeadersList = []
                if firstWordInRow == "{BlockType:List}":
                    isListSection = "Y"
                elif firstWordInRow == "{ReportBlock:End}" and isListSection == 'Y':
                    isListSection = "N"
                continue;
        

        if isListSection == "Y":
            continue;
        ## Need to fix the logic for identifying the Column Start Index 
        ## Presently its hardcoded to 2
        #Check if Row is a Column Header
        columnHeaders = []
        isColHeaderPresent = getColumnHeaders(eachRow, columnHeadersList, colStartidx)
        if isColHeaderPresent == 'Y':
            #print("came here")
            if isColContinous == 'N':    ## reset ColumnHeaders if new list of Column Headers
                #print("came here 2")
                columnHeadersList = [] 
                isResetColStrt = "Y"
            columnHeadersList.append(eachRow)
            isColContinous = 'Y'
        else:
            isColContinous = 'N'
        
        
        if isColHeaderPresent != 'Y':
            #print("Identifying Line Items")
            rowDesc = ""
            for eachCell in range(len(eachRow)):
                
                #print("Col ", eachCell, " is ", eachRow[eachCell], " with color ", eachRowColour[eachCell])
                if eachRow[eachCell] is not None:
                    rowDesc = rowDesc + eachRow[eachCell]
                    #print("RowDesc is ", rowDesc)
                
                elif eachRow[eachCell] is None and eachRowColour[eachCell] == '00000000':
                    #print("Almost there ", len(columnHeadersList))
                    columnHeaders = ""
                    for eachCol in range(len(columnHeadersList)):
                        columnHeaders = columnHeaders + " " + nvl(columnHeadersList[eachCol][eachCell], "")
                    #print("ColHeader is ", columnHeaders)
                    #print("ColHeader length", len(columnHeaders))
                    #print("ColHeader Split length", len(columnHeaders.split()))
                    if len(columnHeaders.split()) > 0 and len(rowDesc.split()) > 0:
                        lineItemDesc = formatStringForLIDesc(rowDesc) + formatStringForLIDesc(columnHeaders)
                        if lineItemDesc.lower().find("total") == -1:
                            ws.cell(row = rowidx, column = eachCell+1).value = "&&LI&&"
                            ws.cell(row = rowidx, column = eachCell+1).fill = PatternFill(start_color= 'FFFFFF00', end_color='FFFFFF00', fill_type= "solid")
                        else:
                            ws.cell(row = rowidx, column = eachCell+1).value = "&&DLI&&"
                            ws.cell(row = rowidx, column = eachCell+1).fill = PatternFill(start_color= 'FF0070C0', end_color='FF0070C0', fill_type= "solid")
                                           
                        if isResetColStrt == "Y":
                            isResetColStrt = "N"
                            colStartidx = eachCell

        #rowidx+=1
        isColHeaderPresent = "N"
        isSectionHeader = "N"
#print("processing done") 

wb.save(r'C:\Users\i31927\Desktop\MLProject\ML-ADDIN\test.xlsm')
#print("all done")   


# %%



