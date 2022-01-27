# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
def getColumnHeaders(row_detail, prev_Col_Header, colStart):
    #print(len(row_detail)," ", colStart)
    for i in range(len(row_detail)):
        if i >= colStart:
            #print(row_detail[i])
            if row_detail[i] is not None:
                #print("this row is a column bcos of ", row_detail[i].split())
                return "Y"
    return "N"

#print("saving done")


# %%
# Initial Configuration
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np 
import win32com.client
from win32com.client import Dispatch


#filename = "Book1.xlsx"
#path     = "D:/Akhil/Work/DOTS/RegAutomation/Akhil/input/20210621"
mode     = "N"  # N(new), M (modified)

#print("After specifying File details")


# %%
# Initialize Variables
rowidx = 0
colidx = 0
colStartidx = 2
columnHeaders = []
columnHeadersList = []
isColHeaderPresent = 'N'
reportBlockList = []
lastColRow = 0
nextRow = 0
reportListBlock = []
offset = 0
eachRow = []
isResetColStrt = "N"
isColContinous = "N"
referenceSheetNames = []
sheetName = ""
#print("after initializing variables")


# %%
#Generate Preview File --> With LineItems and Derived Line Items
#print(path+"/Inp-"+filename)
if mode == "N":
#    wb = openpyxl.load_workbook(path+"/Inp-"+filename)
#    ws = wb.worksheets[0]
    xl = win32com.client.Dispatch('Excel.Application')
    wb = xl.Workbooks.Open(r'C:\Users\i31927\Desktop\MLProject\ML-ADDIN\test.xlsm')
  #  ws = wb.Worksheets('INPUT_DRAFT')
          #       ws.iter_rowsnrows = ws.sheet.UsedRange.Rows.Count # maximum line
          #       ws.ncols = ws.sheet.UsedRange.Columns.Count # maximum column

# Write on empty cell of active sheet
writeData = wb.Worksheets('INPUT_DRAFT')
writeData.Cells(1,1).Value = 'Akhil'

wb.save(r'C:\Users\i31927\Desktop\MLProject\ML-ADDIN\test.xlsm')
print("processing done")

# %%



